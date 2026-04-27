from flask import Flask, render_template, jsonify, request
from apscheduler.schedulers.background import BackgroundScheduler
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, io, json, logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
app = Flask(__name__)

GMAIL_USER       = os.environ.get("GMAIL_USER", "")
GMAIL_PASSWORD   = os.environ.get("GMAIL_PASSWORD", "")
RECIPIENT_EMAIL  = os.environ.get("RECIPIENT_EMAIL", "Muros8478@gmail.com")
SEND_EVERY_HOURS = int(os.environ.get("SEND_EVERY_HOURS", "7"))
KEYWORDS_FILE    = "keywords.json"

def load_keywords():
    try:
        if os.path.exists(KEYWORDS_FILE):
            with open(KEYWORDS_FILE) as f:
                return json.load(f)
    except Exception:
        pass
    return []

def save_keywords(kws):
    try:
        with open(KEYWORDS_FILE, "w") as f:
            json.dump(kws, f)
    except Exception as e:
        logger.error(f"Error keywords: {e}")

try:
    from agente import obtener_convocatorias
    logger.info("Agente OK")
except Exception as e:
    logger.error(f"Error agente: {e}")
    def obtener_convocatorias(palabras_clave=None, tipo_busqueda="ambos"):
        return [], "Sin datos", ""


def build_excel(contratos, tab_nombre="Convocatorias"):
    wb = Workbook()
    ws = wb.active
    ws.title = tab_nombre[:31]
    ROJO, BLANCO, GRIS = "B91C1C", "FFFFFF", "F9FAFB"
    thin   = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"CONVOCATORIAS VIGENTES SEACE/PLADICOP PERU  -  {datetime.now().strftime('%d/%m/%Y %H:%M')}  -  UIT 2026 = S/. 5,500"
    t.font = Font(name="Calibri", bold=True, size=12, color=BLANCO)
    t.fill = PatternFill("solid", fgColor=ROJO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:J2")
    s = ws["A2"]
    s.value = f"Contrato menor maximo: S/. 44,000 (8 UIT)  |  Agente: Google Gemini  |  Total: {len(contratos)} convocatorias"
    s.font = Font(name="Calibri", italic=True, size=10, color="6B7280")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    hdrs   = ["#", "N ORDEN", "TIPO", "OBJETO / BIEN O SERVICIO", "ENTIDAD", "LUGAR", "MONTO S/.", "PRECIO PROM HIST", "RECOMENDACION COTIZACION", "DOCUMENTOS"]
    widths = [4,   13,        12,     42,                          35,        18,       12,           16,                  42,                          38]

    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color=BLANCO)
        c.fill = PatternFill("solid", fgColor=ROJO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 22

    for idx, c in enumerate(contratos, 1):
        fila = idx + 3
        bg   = PatternFill("solid", fgColor=BLANCO if idx % 2 else GRIS)
        try:
            mv = float(str(c.get("monto","0")).replace(",","").replace("S/.","").strip() or 0)
            ms = f"{mv:,.2f}"
        except Exception:
            ms = str(c.get("monto","0"))

        ph = c.get("precio_historico") or {}
        try:
            pp = f"S/. {float(str(ph.get('precio_promedio','0')).replace(',','').strip() or 0):,.2f}"
        except Exception:
            pp = str(ph.get("precio_promedio","-"))

        docs = "\n".join(
            f"- {d.get('nombre','')}: {d.get('url','')}" if d.get("url") else f"- {d.get('nombre','')}"
            for d in (c.get("documentos") or [])
        ) or "Sin documentos"

        vals = [idx, c.get("numero",""), c.get("tipo_contrato",""), c.get("objeto",""),
                c.get("entidad",""), c.get("lugar","-"), ms, pp,
                ph.get("recomendacion",""), docs]
        alns = ["center","center","center","left","left","left","right","right","left","left"]

        for col_i, (val, aln) in enumerate(zip(vals, alns), 1):
            cell = ws.cell(row=fila, column=col_i, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.fill = bg
            cell.alignment = Alignment(horizontal=aln, vertical="top", wrap_text=True)
            cell.border = border

        # Color monto
        try:
            if mv > 0:
                ws.cell(row=fila, column=7).font = Font(name="Calibri", size=9, bold=True, color="065F46")
                ws.cell(row=fila, column=7).fill = PatternFill("solid", fgColor="D1FAE5")
        except Exception:
            pass

        # Color recomendacion
        es_comp = ph.get("es_competitivo")
        rec_cell = ws.cell(row=fila, column=9)
        if es_comp is True:
            rec_cell.fill = PatternFill("solid", fgColor="D1FAE5")
        elif es_comp is False:
            rec_cell.fill = PatternFill("solid", fgColor="FEF3C7")

        ws.row_dimensions[fila].height = max(28, 13 * max(1, len(c.get("documentos") or [])))

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:J{3 + len(contratos)}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_excel_completo(todos, bienes, servicios):
    """Excel con 3 hojas: Todo, Bienes, Servicios."""
    wb = Workbook()
    ROJO, BLANCO, GRIS = "B91C1C", "FFFFFF", "F9FAFB"
    thin   = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdrs   = ["#", "N ORDEN", "TIPO", "OBJETO", "ENTIDAD", "LUGAR", "MONTO S/.", "PRECIO PROM HIST", "RECOMENDACION", "DOCUMENTOS"]
    widths = [4,   13,        12,     42,        35,        18,       12,           16,                  42,              38]

    def fill_sheet(ws, contratos, titulo):
        ws.merge_cells("A1:J1")
        t = ws["A1"]
        t.value = f"{titulo}  -  {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(contratos)}"
        t.font = Font(name="Calibri", bold=True, size=12, color=BLANCO)
        t.fill = PatternFill("solid", fgColor=ROJO)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        for i, (h, w) in enumerate(zip(hdrs, widths), 1):
            c = ws.cell(row=2, column=i, value=h)
            c.font = Font(name="Calibri", bold=True, size=10, color=BLANCO)
            c.fill = PatternFill("solid", fgColor=ROJO)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[2].height = 20

        for idx, c in enumerate(contratos, 1):
            fila = idx + 2
            bg   = PatternFill("solid", fgColor=BLANCO if idx % 2 else GRIS)
            try:
                mv = float(str(c.get("monto","0")).replace(",","").replace("S/.","").strip() or 0)
                ms = f"{mv:,.2f}"
            except Exception:
                ms = str(c.get("monto","0"))
            ph = c.get("precio_historico") or {}
            try:
                pp = f"S/. {float(str(ph.get('precio_promedio','0')).replace(',','').strip() or 0):,.2f}"
            except Exception:
                pp = "-"
            docs = "; ".join(d.get("nombre","") for d in (c.get("documentos") or [])) or "Sin documentos"
            vals = [idx, c.get("numero",""), c.get("tipo_contrato",""), c.get("objeto",""),
                    c.get("entidad",""), c.get("lugar","-"), ms, pp,
                    ph.get("recomendacion",""), docs]
            alns = ["center","center","center","left","left","left","right","right","left","left"]
            for col_i, (val, aln) in enumerate(zip(vals, alns), 1):
                cell = ws.cell(row=fila, column=col_i, value=val)
                cell.font = Font(name="Calibri", size=9)
                cell.fill = bg
                cell.alignment = Alignment(horizontal=aln, vertical="top", wrap_text=True)
                cell.border = border
            try:
                if mv > 0:
                    ws.cell(row=fila, column=7).font = Font(name="Calibri", size=9, bold=True, color="065F46")
                    ws.cell(row=fila, column=7).fill = PatternFill("solid", fgColor="D1FAE5")
            except Exception:
                pass
            ws.row_dimensions[fila].height = 26

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:J{2 + len(contratos)}"

    # Hoja 1: Todos
    ws1 = wb.active
    ws1.title = "Todas las Convocatorias"
    fill_sheet(ws1, todos, "TODAS LAS CONVOCATORIAS VIGENTES - SEACE/PLADICOP PERU")

    # Hoja 2: Bienes
    ws2 = wb.create_sheet("Bienes")
    fill_sheet(ws2, bienes, "CONVOCATORIAS DE BIENES VIGENTES")

    # Hoja 3: Servicios
    ws3 = wb.create_sheet("Servicios")
    fill_sheet(ws3, servicios, "CONVOCATORIAS DE SERVICIOS VIGENTES")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_html_email(todos, bienes, servicios, fuente, periodo, keywords):
    hoy  = datetime.now().strftime("%d/%m/%Y")
    hora = datetime.now().strftime("%H:%M")
    kw_txt = ", ".join(keywords) if keywords else "Todos"

    def filas_html(contratos):
        html = ""
        for i, c in enumerate(contratos):
            bg = "#ffffff" if i % 2 == 0 else "#f9fafb"
            try:
                mv = float(str(c.get("monto","0")).replace(",","").replace("S/.","").strip() or 0)
                ms = f"S/. {mv:,.2f}" if mv > 0 else "-"
            except Exception:
                ms = "-"
            ph = c.get("precio_historico") or {}
            try:
                pp = f"S/. {float(str(ph.get('precio_promedio','0')).replace(',','').strip() or 0):,.2f}"
            except Exception:
                pp = "-"
            rec = ph.get("recomendacion","")
            es_comp = ph.get("es_competitivo")
            rec_bg = "#d1fae5" if es_comp is True else "#fef3c7" if es_comp is False else "#f9fafb"
            tipo_badge = "🏭" if c.get("tipo_contrato","") == "Bienes" else "🔧"
            docs_html = "".join(
                f'<a href="{d.get("url","")}" style="display:inline-block;margin:2px;padding:2px 7px;background:#fef3c7;color:#92400e;border-radius:4px;font-size:10px;font-weight:600;text-decoration:none">📄 {d.get("nombre","")}</a>'
                if d.get("url") else ""
                for d in (c.get("documentos") or [])
            )
            url = c.get("urlSeace","")
            obj = c.get("objeto","")
            obj_html = f'<a href="{url}" style="color:#b91c1c;font-weight:600;text-decoration:none;font-size:13px">{obj}</a>' if url else f'<strong style="font-size:13px">{obj}</strong>'
            html += f"""<tr style="background:{bg}">
              <td style="padding:8px 10px;font-size:12px;border-bottom:1px solid #e5e7eb;text-align:center;vertical-align:top">{i+1}</td>
              <td style="padding:8px 10px;border-bottom:1px solid #e5e7eb;vertical-align:top">
                <div>{tipo_badge} {obj_html}</div>
                <div style="font-size:11px;color:#6b7280;margin-top:2px">{c.get('entidad','')} · 📍{c.get('lugar','')}</div>
                <div style="margin-top:4px">{docs_html}</div>
              </td>
              <td style="padding:8px 10px;font-size:12px;font-weight:700;color:#065f46;border-bottom:1px solid #e5e7eb;text-align:right;vertical-align:top;white-space:nowrap">{ms}</td>
              <td style="padding:8px 10px;font-size:12px;border-bottom:1px solid #e5e7eb;text-align:right;vertical-align:top;white-space:nowrap;color:#374151">{pp}</td>
              <td style="padding:8px 10px;font-size:11px;border-bottom:1px solid #e5e7eb;vertical-align:top;background:{rec_bg}">{rec}</td>
            </tr>"""
        return html

    def seccion(titulo, contratos, color):
        if not contratos:
            return ""
        return f"""
        <div style="margin-bottom:24px">
          <div style="background:{color};padding:10px 16px;border-radius:8px 8px 0 0">
            <h3 style="margin:0;color:#fff;font-size:14px;font-weight:700">{titulo} ({len(contratos)})</h3>
          </div>
          <div style="overflow-x:auto;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 8px 8px">
            <table style="width:100%;border-collapse:collapse">
              <thead><tr style="background:#f9fafb">
                <th style="padding:8px 10px;font-size:11px;color:#6b7280;border-bottom:2px solid #e5e7eb">#</th>
                <th style="padding:8px 10px;font-size:11px;color:#6b7280;border-bottom:2px solid #e5e7eb;text-align:left">OBJETO / ENTIDAD / DOCS</th>
                <th style="padding:8px 10px;font-size:11px;color:#6b7280;border-bottom:2px solid #e5e7eb;text-align:right">MONTO</th>
                <th style="padding:8px 10px;font-size:11px;color:#6b7280;border-bottom:2px solid #e5e7eb;text-align:right">PRECIO HIST.</th>
                <th style="padding:8px 10px;font-size:11px;color:#6b7280;border-bottom:2px solid #e5e7eb;text-align:left">RECOMENDACION</th>
              </tr></thead>
              <tbody>{filas_html(contratos)}</tbody>
            </table>
          </div>
        </div>"""

    return f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:#f3f4f6;font-family:Arial,sans-serif">
<div style="max-width:980px;margin:20px auto;background:#fff;border-radius:12px;overflow:hidden">
  <div style="background:#b91c1c;padding:20px 24px">
    <h1 style="margin:0;color:#fff;font-size:20px;font-weight:800">🇵🇪 Convocatorias Vigentes SEACE/PLADICOP Peru</h1>
    <div style="color:rgba(255,255,255,0.85);font-size:13px;margin-top:6px">
      📅 {hoy} · 🕐 {hora} · 📋 {len(todos)} convocatorias totales ({len(bienes)} bienes + {len(servicios)} servicios)
      · UIT 2026: S/. 5,500 · Maximo contrato menor: S/. 44,000
    </div>
    <div style="color:rgba(255,255,255,0.7);font-size:12px;margin-top:4px">
      Periodo: {periodo} · Rubro filtrado: {kw_txt} · Fuente: {fuente}
    </div>
  </div>
  <div style="padding:20px 24px">
    {seccion("🏭 Bienes — Materiales, Equipos y Productos", bienes, "#065f46")}
    {seccion("🔧 Servicios — Mantenimiento, Limpieza y Otros", servicios, "#1e40af")}
  </div>
  <div style="padding:12px 24px 20px;border-top:1px solid #f3f4f6;background:#fafafa">
    <p style="margin:0;font-size:11px;color:#9ca3af">
      📊 Excel adjunto con 3 hojas: Todas, Bienes, Servicios · 🤖 Google Gemini AI
      · Fuente: <a href="https://prod6.seace.gob.pe/buscador-publico/contrataciones" style="color:#b91c1c">SEACE</a>
      · Envio automatico cada {SEND_EVERY_HOURS}h
    </p>
  </div>
</div></body></html>"""


def send_report(todos, bienes, servicios, fuente, periodo, keywords=None):
    if not GMAIL_USER or not GMAIL_PASSWORD:
        raise ValueError("GMAIL_USER y GMAIL_PASSWORD no configurados.")
    hoy  = datetime.now().strftime("%d/%m/%Y")
    hora = datetime.now().strftime("%H:%M")
    msg  = MIMEMultipart("mixed")
    msg["Subject"] = f"SEACE {hoy} {hora} - {len(todos)} convocatorias ({len(bienes)} bienes + {len(servicios)} servicios)"
    msg["From"]    = GMAIL_USER
    msg["To"]      = RECIPIENT_EMAIL
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(build_html_email(todos, bienes, servicios, fuente, periodo, keywords or []), "html", "utf-8"))
    msg.attach(alt)
    excel_buf = build_excel_completo(todos, bienes, servicios)
    nombre = f"SEACE_Convocatorias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(excel_buf.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{nombre}"')
    msg.attach(part)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_PASSWORD)
        server.sendmail(GMAIL_USER, RECIPIENT_EMAIL, msg.as_string())
    logger.info(f"Correo enviado: {len(todos)} convocatorias")
    return len(todos)


def job_automatico():
    logger.info(f"Job {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    try:
        keywords = load_keywords()
        todos, fuente, periodo = obtener_convocatorias(palabras_clave=keywords or None, tipo_busqueda="ambos")
        bienes   = [c for c in todos if c.get("tipo_contrato","") == "Bienes"]
        servicios= [c for c in todos if c.get("tipo_contrato","") == "Servicios"]
        if todos:
            send_report(todos, bienes, servicios, fuente, periodo, keywords)
    except Exception as e:
        logger.error(f"Error job: {e}")

try:
    scheduler = BackgroundScheduler(daemon=True)
    scheduler.add_job(job_automatico, "interval", hours=SEND_EVERY_HOURS)
    scheduler.start()
    logger.info(f"Scheduler OK: cada {SEND_EVERY_HOURS}h")
except Exception as e:
    logger.error(f"Error scheduler: {e}")


@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/convocatorias")
def api_convocatorias():
    try:
        keywords = load_keywords()
        tipo = request.args.get("tipo", "ambos")
        todos, fuente, periodo = obtener_convocatorias(palabras_clave=keywords or None, tipo_busqueda=tipo)
        bienes    = [c for c in todos if c.get("tipo_contrato","") == "Bienes"]
        servicios = [c for c in todos if c.get("tipo_contrato","") == "Servicios"]
        return jsonify({"ok": True, "total": len(todos), "contratos": todos,
                        "bienes": bienes, "servicios": servicios,
                        "fuente": fuente, "periodo": periodo,
                        "fecha": datetime.now().strftime("%d/%m/%Y %H:%M"),
                        "keywords": keywords})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/enviar", methods=["POST"])
def api_enviar():
    try:
        body      = request.get_json() or {}
        keywords  = load_keywords()
        contratos = body.get("contratos") or []
        if not contratos:
            contratos, fuente, periodo = obtener_convocatorias(tipo_busqueda="ambos")
        else:
            fuente, periodo = "SEACE Peru", ""
        bienes    = [c for c in contratos if c.get("tipo_contrato","") == "Bienes"]
        servicios = [c for c in contratos if c.get("tipo_contrato","") == "Servicios"]
        total = send_report(contratos, bienes, servicios, fuente, periodo, keywords)
        return jsonify({"ok": True, "message": f"Excel enviado a {RECIPIENT_EMAIL} con {total} convocatorias ({len(bienes)} bienes + {len(servicios)} servicios)."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/keywords", methods=["GET"])
def api_get_keywords():
    return jsonify({"keywords": load_keywords()})

@app.route("/api/keywords", methods=["POST"])
def api_save_keywords():
    try:
        body = request.get_json() or {}
        kws  = [k.strip() for k in body.get("keywords", []) if k.strip()]
        save_keywords(kws)
        return jsonify({"ok": True, "keywords": kws})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/config")
def api_config():
    return jsonify({
        "gmail_configurado":  bool(GMAIL_USER and GMAIL_PASSWORD),
        "gemini_configurado": bool(os.environ.get("GEMINI_API_KEY","")),
        "destinatario":       RECIPIENT_EMAIL,
        "frecuencia":         f"Cada {SEND_EVERY_HOURS} horas",
        "uit_2026":           "S/. 5,500",
        "maximo_contrato":    "S/. 44,000",
        "keywords":           load_keywords(),
    })

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
